from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

ARQUIVO = "Controle_Estoque_Tintas_Inflamaveis.xlsx"

wb = Workbook()
ws_painel = wb.active
ws_painel.title = "Painel"
ws_prod = wb.create_sheet("Cadastro_Produtos")
ws_mov = wb.create_sheet("Movimentacoes")
ws_est = wb.create_sheet("Estoque_Validades")
ws_instr = wb.create_sheet("Instrucoes")

# Estilos base
cor_titulo = PatternFill("solid", fgColor="1F4E78")
cor_header = PatternFill("solid", fgColor="2F75B5")
cor_card = PatternFill("solid", fgColor="D9E1F2")
cor_botao = PatternFill("solid", fgColor="00B050")
fonte_titulo = Font(color="FFFFFF", bold=True, size=14)
fonte_header = Font(color="FFFFFF", bold=True)
fonte_botao = Font(color="FFFFFF", bold=True)
fonte_bold = Font(bold=True)
centro = Alignment(horizontal="center", vertical="center")
esquerda = Alignment(horizontal="left", vertical="center")
borda = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def ajustar_larguras(ws, larguras):
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


def aplicar_header(ws, linha, colunas):
    for i, nome in enumerate(colunas, 1):
        c = ws.cell(row=linha, column=i, value=nome)
        c.fill = cor_header
        c.font = fonte_header
        c.alignment = centro
        c.border = borda


def pintar_tabela(ws, ini, fim, cols):
    for r in range(ini, fim + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = borda
            if c in (1, 2, 3):
                cell.alignment = centro
            else:
                cell.alignment = esquerda


# ---------------- Painel ----------------
ws_painel.merge_cells("A1:J1")
ws_painel["A1"] = "CONTROLE DE ESTOQUE - TINTAS E INFLAMAVEIS"
ws_painel["A1"].fill = cor_titulo
ws_painel["A1"].font = fonte_titulo
ws_painel["A1"].alignment = centro

cards = [
    ("A3:C5", "Produtos Cadastrados", "=COUNTA(Cadastro_Produtos!A6:A2000)"),
    ("D3:F5", "Lotes Cadastrados", "=COUNTA(Estoque_Validades!C6:C5000)"),
    ("G3:J5", "Saldo Total (Unid.)", "=SUM(Estoque_Validades!J6:J5000)"),
    ("A6:C8", "Vencidos", '=COUNTIF(Estoque_Validades!F6:F5000,"VENCIDO")'),
    ("D6:F8", "Vence em ate 30 dias", '=COUNTIF(Estoque_Validades!F6:F5000,"VENCE EM <= 30 DIAS")'),
    ("G6:J8", "Abaixo do Estoque Minimo", '=COUNTIF(Estoque_Validades!K6:K5000,"ABAIXO DO MINIMO")'),
]

for area, titulo, formula in cards:
    ws_painel.merge_cells(area)
    c = ws_painel[area.split(":")[0]]
    c.value = f"{titulo}\n{formula}"
    c.fill = cor_card
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Botões (células clicáveis)
botoes = [
    ("A10:C11", "Ir para Cadastro de Produtos", "#'Cadastro_Produtos'!A1"),
    ("D10:F11", "Ir para Movimentacoes", "#'Movimentacoes'!A1"),
    ("G10:J11", "Ir para Estoque e Validades", "#'Estoque_Validades'!A1"),
    ("A12:J13", "Ir para Instrucoes de Uso", "#'Instrucoes'!A1"),
]
for area, texto, link in botoes:
    ws_painel.merge_cells(area)
    c = ws_painel[area.split(":")[0]]
    c.value = texto
    c.hyperlink = link
    c.fill = cor_botao
    c.font = fonte_botao
    c.alignment = centro
    for row in ws_painel[area]:
        for cell in row:
            cell.border = borda

ajustar_larguras(ws_painel, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18})

# ---------------- Cadastro de Produtos ----------------
ws_prod.merge_cells("A1:G1")
ws_prod["A1"] = "CADASTRO DE PRODUTOS"
ws_prod["A1"].fill = cor_titulo
ws_prod["A1"].font = fonte_titulo
ws_prod["A1"].alignment = centro

ws_prod["A3"] = "Preencha os produtos a partir da linha 6."
ws_prod["A3"].font = fonte_bold

headers_prod = [
    "ID_Produto", "Produto", "Categoria", "Unidade", "Estoque_Minimo", "Local_Armazenagem", "Observacoes"
]
aplicar_header(ws_prod, 5, headers_prod)

# Exemplo inicial
ws_prod["A6"] = "PRD001"
ws_prod["B6"] = "Tinta Epoxi Azul"
ws_prod["C6"] = "TINTA"
ws_prod["D6"] = "L"
ws_prod["E6"] = 20
ws_prod["F6"] = "Dep. A - Prateleira 1"
ws_prod["G6"] = "Exemplo"

# Data validation categoria/unidade
dv_categoria = DataValidation(type="list", formula1='"TINTA,INFLAMAVEL"', allow_blank=True)
dv_unidade = DataValidation(type="list", formula1='"L,KG,UN"', allow_blank=True)
ws_prod.add_data_validation(dv_categoria)
ws_prod.add_data_validation(dv_unidade)
dv_categoria.add("C6:C2000")
dv_unidade.add("D6:D2000")

pintar_tabela(ws_prod, 6, 200, 7)
ajustar_larguras(ws_prod, {"A": 14, "B": 30, "C": 14, "D": 10, "E": 16, "F": 24, "G": 30})

# ---------------- Movimentações ----------------
ws_mov.merge_cells("A1:J1")
ws_mov["A1"] = "LANCAMENTO DE MOVIMENTACOES"
ws_mov["A1"].fill = cor_titulo
ws_mov["A1"].font = fonte_titulo
ws_mov["A1"].alignment = centro

ws_mov["A3"] = "Registre ENTRADA, SAIDA_USO e SAIDA_DESCARTE a partir da linha 6."
ws_mov["A3"].font = fonte_bold

headers_mov = [
    "Data", "Tipo", "ID_Produto", "Produto", "Lote", "Validade", "Quantidade", "Responsavel", "Destino_ou_Motivo", "Observacoes"
]
aplicar_header(ws_mov, 5, headers_mov)

# Linha exemplo
ws_mov["A6"] = "10/03/2026"
ws_mov["B6"] = "ENTRADA"
ws_mov["C6"] = "PRD001"
ws_mov["D6"] = '=IFERROR(VLOOKUP(C6,Cadastro_Produtos!A:B,2,FALSE),"")'
ws_mov["E6"] = "LOT-001"
ws_mov["F6"] = "20/06/2026"
ws_mov["G6"] = 50
ws_mov["H6"] = "Almoxarife"
ws_mov["I6"] = "Entrada inicial"
ws_mov["J6"] = "Exemplo"

# Fórmulas de produto automática
for r in range(7, 5001):
    ws_mov[f"D{r}"] = f'=IFERROR(VLOOKUP(C{r},Cadastro_Produtos!A:B,2,FALSE),"")'

# Validações
dv_tipo = DataValidation(type="list", formula1='"ENTRADA,SAIDA_USO,SAIDA_DESCARTE"', allow_blank=False)
ws_mov.add_data_validation(dv_tipo)
dv_tipo.add("B6:B5000")

for col in ["A", "F"]:
    for r in range(6, 5001):
        ws_mov[f"{col}{r}"].number_format = "dd/mm/yyyy"
ws_mov["G6"].number_format = "0.00"
for r in range(7, 5001):
    ws_mov[f"G{r}"].number_format = "0.00"

pintar_tabela(ws_mov, 6, 220, 10)
ajustar_larguras(ws_mov, {
    "A": 12, "B": 17, "C": 14, "D": 28, "E": 14, "F": 12, "G": 12, "H": 18, "I": 24, "J": 26
})

# ---------------- Estoque e Validades ----------------
ws_est.merge_cells("A1:K1")
ws_est["A1"] = "ESTOQUE, ALERTAS E VALIDADES POR LOTE"
ws_est["A1"].fill = cor_titulo
ws_est["A1"].font = fonte_titulo
ws_est["A1"].alignment = centro

ws_est["A3"] = "Cadastre cada lote uma vez e o saldo sera calculado automaticamente pelas movimentacoes."
ws_est["A3"].font = fonte_bold

headers_est = [
    "ID_Produto", "Produto", "Lote", "Validade", "Dias_para_Vencer", "Status_Validade",
    "Entradas", "Saida_Uso", "Saida_Descarte", "Saldo", "Alerta_Estoque"
]
aplicar_header(ws_est, 5, headers_est)

# Exemplo + fórmulas
ws_est["A6"] = "PRD001"
ws_est["B6"] = '=IFERROR(VLOOKUP(A6,Cadastro_Produtos!A:B,2,FALSE),"")'
ws_est["C6"] = "LOT-001"
ws_est["D6"] = "20/06/2026"

for r in range(6, 5001):
    ws_est[f"B{r}"] = f'=IFERROR(VLOOKUP(A{r},Cadastro_Produtos!A:B,2,FALSE),"")'
    ws_est[f"E{r}"] = f'=IF(D{r}="","",D{r}-TODAY())'
    ws_est[f"F{r}"] = (
        f'=IF(D{r}="","",IF(E{r}<0,"VENCIDO",IF(E{r}<=30,"VENCE EM <= 30 DIAS",IF(E{r}<=60,"ATENCAO (31-60 DIAS)","OK"))))'
    )
    ws_est[f"G{r}"] = f'=SUMIFS(Movimentacoes!$G:$G,Movimentacoes!$C:$C,$A{r},Movimentacoes!$E:$E,$C{r},Movimentacoes!$B:$B,"ENTRADA")'
    ws_est[f"H{r}"] = f'=SUMIFS(Movimentacoes!$G:$G,Movimentacoes!$C:$C,$A{r},Movimentacoes!$E:$E,$C{r},Movimentacoes!$B:$B,"SAIDA_USO")'
    ws_est[f"I{r}"] = f'=SUMIFS(Movimentacoes!$G:$G,Movimentacoes!$C:$C,$A{r},Movimentacoes!$E:$E,$C{r},Movimentacoes!$B:$B,"SAIDA_DESCARTE")'
    ws_est[f"J{r}"] = f'=G{r}-H{r}-I{r}'
    ws_est[f"K{r}"] = f'=IF(A{r}="","",IF(J{r}<=IFERROR(VLOOKUP(A{r},Cadastro_Produtos!$A:$E,5,FALSE),0),"ABAIXO DO MINIMO","OK"))'

for col in ["D"]:
    for r in range(6, 5001):
        ws_est[f"{col}{r}"].number_format = "dd/mm/yyyy"
for col in ["E", "G", "H", "I", "J"]:
    for r in range(6, 5001):
        ws_est[f"{col}{r}"].number_format = "0.00"

# Formatação condicional status validade
regra_vencido = FormulaRule(formula=['$F6="VENCIDO"'], fill=PatternFill("solid", fgColor="FFC7CE"))
regra_30 = FormulaRule(formula=['$F6="VENCE EM <= 30 DIAS"'], fill=PatternFill("solid", fgColor="FFEB9C"))
regra_60 = FormulaRule(formula=['$F6="ATENCAO (31-60 DIAS)"'], fill=PatternFill("solid", fgColor="FCE4D6"))
ws_est.conditional_formatting.add("A6:K5000", regra_vencido)
ws_est.conditional_formatting.add("A6:K5000", regra_30)
ws_est.conditional_formatting.add("A6:K5000", regra_60)

# Formatação condicional estoque mínimo
regra_minimo = FormulaRule(formula=['$K6="ABAIXO DO MINIMO"'], fill=PatternFill("solid", fgColor="F4B183"))
ws_est.conditional_formatting.add("A6:K5000", regra_minimo)

pintar_tabela(ws_est, 6, 220, 11)
ajustar_larguras(ws_est, {
    "A": 14, "B": 28, "C": 14, "D": 12, "E": 15, "F": 24,
    "G": 12, "H": 12, "I": 14, "J": 12, "K": 18
})

# ---------------- Instruções ----------------
ws_instr.merge_cells("A1:H1")
ws_instr["A1"] = "INSTRUCOES DE USO"
ws_instr["A1"].fill = cor_titulo
ws_instr["A1"].font = fonte_titulo
ws_instr["A1"].alignment = centro

instrucoes = [
    "1) Aba Cadastro_Produtos: cadastre todos os itens (ID unico, nome, categoria e estoque minimo).",
    "2) Aba Estoque_Validades: cadastre cada lote (ID produto + lote + validade), uma linha por lote.",
    "3) Aba Movimentacoes: registre toda ENTRADA, SAIDA_USO e SAIDA_DESCARTE com data, lote e quantidade.",
    "4) O saldo por lote e os alertas sao automaticos na aba Estoque_Validades.",
    "5) Alertas de validade: VENCIDO (vermelho), VENCE EM <= 30 DIAS (amarelo), ATENCAO 31-60 dias (laranja claro).",
    "6) Alerta de estoque minimo: coluna Alerta_Estoque marca ABAIXO DO MINIMO.",
    "7) Aba Painel: use os botoes verdes para navegar e acompanhar os indicadores gerais.",
    "8) Recomendacao operacional: registre a movimentacao no mesmo dia para manter o saldo correto.",
    "9) Para novos produtos, sempre crie o cadastro antes de lancar lote/movimentacao.",
    "10) Nao altere formulas nas colunas calculadas (Produto, Dias, Status, Entradas, Saidas, Saldo e Alertas).",
]

ws_instr["A3"] = "Fluxo recomendado"
ws_instr["A3"].font = fonte_bold
for i, txt in enumerate(instrucoes, start=5):
    ws_instr[f"A{i}"] = txt
    ws_instr[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")

ws_instr["A18"] = "Dica: filtre as tabelas por status para priorizar produtos vencidos e proximos do vencimento."
ws_instr["A18"].font = Font(italic=True)

ajustar_larguras(ws_instr, {"A": 120, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})

# Congelar painéis
ws_prod.freeze_panes = "A6"
ws_mov.freeze_panes = "A6"
ws_est.freeze_panes = "A6"

# AutoFilter
ws_prod.auto_filter.ref = "A5:G2000"
ws_mov.auto_filter.ref = "A5:J5000"
ws_est.auto_filter.ref = "A5:K5000"

wb.save(ARQUIVO)
print(f"Arquivo criado: {ARQUIVO}")
