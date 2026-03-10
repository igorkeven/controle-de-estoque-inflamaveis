from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

ARQ = "Sistema_Estoque_Inflamaveis_SEM_INSTALACAO.xlsx"

wb = Workbook()
ws_home = wb.active
ws_home.title = "Inicio"
ws_prod = wb.create_sheet("Produtos")
ws_lotes = wb.create_sheet("Lotes")
ws_mov = wb.create_sheet("Lancamentos")
ws_stock = wb.create_sheet("Estoque_Alertas")
ws_rel = wb.create_sheet("Relatorio")
ws_help = wb.create_sheet("Como_Usar")

# estilos
c_titulo = PatternFill("solid", fgColor="0E3A52")
c_header = PatternFill("solid", fgColor="145A7B")
c_box = PatternFill("solid", fgColor="EAF4FB")
c_btn = PatternFill("solid", fgColor="1F9D67")
c_input = PatternFill("solid", fgColor="FFF8E7")

f_titulo = Font(color="FFFFFF", bold=True, size=14)
f_head = Font(color="FFFFFF", bold=True)
f_btn = Font(color="FFFFFF", bold=True)
f_b = Font(bold=True)

ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")

b = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def colw(ws, data):
    for col, w in data.items():
        ws.column_dimensions[col].width = w

def head(ws, row, names):
    for i, n in enumerate(names, 1):
        c = ws.cell(row=row, column=i, value=n)
        c.fill = c_header
        c.font = f_head
        c.alignment = ctr
        c.border = b

def table_border(ws, r1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = b
            cell.alignment = left if c not in (1, 2, 3, 4, 5) else ctr

# INICIO
ws_home.merge_cells("A1:J1")
ws_home["A1"] = "SISTEMA DE CONTROLE DE ESTOQUE - TINTAS E INFLAMAVEIS"
ws_home["A1"].fill = c_titulo
ws_home["A1"].font = f_titulo
ws_home["A1"].alignment = ctr

cards = [
    ("A3:C5", "Produtos", "=COUNTA(Produtos!A7:A3000)"),
    ("D3:F5", "Lotes", "=COUNTA(Lotes!C7:C8000)"),
    ("G3:J5", "Saldo Total", "=SUM(Estoque_Alertas!J7:J8000)"),
    ("A6:C8", "Vencidos", '=COUNTIF(Estoque_Alertas!F7:F8000,"VENCIDO")'),
    ("D6:F8", "Vence <=30 dias", '=COUNTIF(Estoque_Alertas!F7:F8000,"VENCE EM <= 30 DIAS")'),
    ("G6:J8", "Abaixo Minimo", '=COUNTIF(Estoque_Alertas!K7:K8000,"ABAIXO DO MINIMO")'),
]
for area, t, f in cards:
    ws_home.merge_cells(area)
    c = ws_home[area.split(":")[0]]
    c.value = f"{t}\n{f}"
    c.fill = c_box
    c.alignment = ctr
    c.font = Font(size=11, bold=True)

btns = [
    ("A10:C11", "1) Cadastrar Produtos", "#'Produtos'!A1"),
    ("D10:F11", "2) Cadastrar Lotes", "#'Lotes'!A1"),
    ("G10:J11", "3) Lancar Movimentacoes", "#'Lancamentos'!A1"),
    ("A12:E13", "4) Ver Estoque e Alertas", "#'Estoque_Alertas'!A1"),
    ("F12:J13", "5) Ler Instrucoes", "#'Como_Usar'!A1"),
]
for area, txt, link in btns:
    ws_home.merge_cells(area)
    c = ws_home[area.split(":")[0]]
    c.value = txt
    c.hyperlink = link
    c.fill = c_btn
    c.font = f_btn
    c.alignment = ctr
    for row in ws_home[area]:
        for cell in row:
            cell.border = b

colw(ws_home, {k: 17 for k in "ABCDEFGHIJ"})

# PRODUTOS
ws_prod.merge_cells("A1:G1")
ws_prod["A1"] = "PASSO 1 - CADASTRO DE PRODUTOS"
ws_prod["A1"].fill = c_titulo
ws_prod["A1"].font = f_titulo
ws_prod["A1"].alignment = ctr
ws_prod["A3"] = "Preencha uma linha por produto a partir da linha 7."
ws_prod["A3"].font = f_b

head(ws_prod, 6, ["ID_Produto", "Nome", "Categoria", "Unidade", "Estoque_Minimo", "Local", "Observacoes"])

ws_prod["A7"] = "PRD001"
ws_prod["B7"] = "Tinta Poliuretano Branca"
ws_prod["C7"] = "TINTA"
ws_prod["D7"] = "L"
ws_prod["E7"] = 20
ws_prod["F7"] = "Deposito A"
ws_prod["G7"] = "Exemplo"

dv_cat = DataValidation(type="list", formula1='"TINTA,INFLAMAVEL"', allow_blank=True)
dv_uni = DataValidation(type="list", formula1='"L,KG,UN"', allow_blank=True)
ws_prod.add_data_validation(dv_cat)
ws_prod.add_data_validation(dv_uni)
dv_cat.add("C7:C3000")
dv_uni.add("D7:D3000")

table_border(ws_prod, 7, 260, 7)
ws_prod.freeze_panes = "A7"
ws_prod.auto_filter.ref = "A6:G3000"
colw(ws_prod, {"A":14,"B":30,"C":14,"D":10,"E":16,"F":20,"G":26})

# LOTES
ws_lotes.merge_cells("A1:H1")
ws_lotes["A1"] = "PASSO 2 - CADASTRO DE LOTES"
ws_lotes["A1"].fill = c_titulo
ws_lotes["A1"].font = f_titulo
ws_lotes["A1"].alignment = ctr
ws_lotes["A3"] = "Cada lote deve ser cadastrado uma unica vez (ID_Produto + Lote + Validade)."
ws_lotes["A3"].font = f_b

head(ws_lotes, 6, ["ID_Produto", "Produto", "Lote", "Fabricacao", "Validade", "Fornecedor", "Nota_Fiscal", "Obs"])

ws_lotes["A7"] = "PRD001"
ws_lotes["B7"] = '=IFERROR(VLOOKUP(A7,Produtos!A:B,2,FALSE),"")'
ws_lotes["C7"] = "LOT001"
ws_lotes["D7"] = "01/01/2026"
ws_lotes["E7"] = "30/06/2026"
ws_lotes["F7"] = "Fornecedor Exemplo"
ws_lotes["G7"] = "NF0001"
ws_lotes["H7"] = "Exemplo"

for r in range(8, 8001):
    ws_lotes[f"B{r}"] = f'=IFERROR(VLOOKUP(A{r},Produtos!A:B,2,FALSE),"")'

for c in ["D", "E"]:
    for r in range(7, 8001):
        ws_lotes[f"{c}{r}"].number_format = "dd/mm/yyyy"

table_border(ws_lotes, 7, 260, 8)
ws_lotes.freeze_panes = "A7"
ws_lotes.auto_filter.ref = "A6:H8000"
colw(ws_lotes, {"A":14,"B":28,"C":14,"D":12,"E":12,"F":20,"G":14,"H":24})

# LANCAMENTOS
ws_mov.merge_cells("A1:J1")
ws_mov["A1"] = "PASSO 3 - LANCAMENTOS DE ENTRADA E SAIDAS"
ws_mov["A1"].fill = c_titulo
ws_mov["A1"].font = f_titulo
ws_mov["A1"].alignment = ctr

ws_mov["A3"] = "Registre ENTRADA, SAIDA_USO e SAIDA_DESCARTE."
ws_mov["A3"].font = f_b

# área de entrada destacada
ws_mov.merge_cells("A4:J4")
ws_mov["A4"] = "AREA DE DIGITACAO (linha 7 em diante)"
ws_mov["A4"].fill = c_input
ws_mov["A4"].font = f_b
ws_mov["A4"].alignment = ctr

head(ws_mov, 6, ["Data", "Tipo", "ID_Produto", "Produto", "Lote", "Validade", "Quantidade", "Responsavel", "Destino/Motivo", "Obs"])

ws_mov["A7"] = "10/03/2026"
ws_mov["B7"] = "ENTRADA"
ws_mov["C7"] = "PRD001"
ws_mov["D7"] = '=IFERROR(VLOOKUP(C7,Produtos!A:B,2,FALSE),"")'
ws_mov["E7"] = "LOT001"
ws_mov["F7"] = '=IFERROR(INDEX(Lotes!$E:$E,MATCH(1,(Lotes!$A:$A=C7)*(Lotes!$C:$C=E7),0)),"")'
ws_mov["G7"] = 50
ws_mov["H7"] = "Almoxarife"
ws_mov["I7"] = "Entrada inicial"
ws_mov["J7"] = "Exemplo"

for r in range(8, 12001):
    ws_mov[f"D{r}"] = f'=IFERROR(VLOOKUP(C{r},Produtos!A:B,2,FALSE),"")'
    ws_mov[f"F{r}"] = f'=IFERROR(INDEX(Lotes!$E:$E,MATCH(1,(Lotes!$A:$A=C{r})*(Lotes!$C:$C=E{r}),0)),"")'

for col in ["A", "F"]:
    for r in range(7, 12001):
        ws_mov[f"{col}{r}"].number_format = "dd/mm/yyyy"
for r in range(7, 12001):
    ws_mov[f"G{r}"].number_format = "0.00"

dv_tipo = DataValidation(type="list", formula1='"ENTRADA,SAIDA_USO,SAIDA_DESCARTE"', allow_blank=False)
ws_mov.add_data_validation(dv_tipo)
dv_tipo.add("B7:B12000")

table_border(ws_mov, 7, 300, 10)
ws_mov.freeze_panes = "A7"
ws_mov.auto_filter.ref = "A6:J12000"
colw(ws_mov, {"A":12,"B":17,"C":14,"D":26,"E":14,"F":12,"G":12,"H":16,"I":22,"J":24})

# ESTOQUE E ALERTAS
ws_stock.merge_cells("A1:K1")
ws_stock["A1"] = "PASSO 4 - ESTOQUE E ALERTAS AUTOMATICOS"
ws_stock["A1"].fill = c_titulo
ws_stock["A1"].font = f_titulo
ws_stock["A1"].alignment = ctr

ws_stock["A3"] = "Nao editar colunas calculadas. Cadastre somente ID_Produto e Lote quando for novo lote."
ws_stock["A3"].font = f_b

head(ws_stock, 6, ["ID_Produto", "Produto", "Lote", "Validade", "Dias", "Status_Validade", "Entradas", "Saida_Uso", "Saida_Descarte", "Saldo", "Alerta_Estoque"])

ws_stock["A7"] = "PRD001"
ws_stock["B7"] = '=IFERROR(VLOOKUP(A7,Produtos!A:B,2,FALSE),"")'
ws_stock["C7"] = "LOT001"
ws_stock["D7"] = '=IFERROR(INDEX(Lotes!$E:$E,MATCH(1,(Lotes!$A:$A=A7)*(Lotes!$C:$C=C7),0)),"")'

for r in range(7, 8001):
    ws_stock[f"B{r}"] = f'=IFERROR(VLOOKUP(A{r},Produtos!A:B,2,FALSE),"")'
    ws_stock[f"D{r}"] = f'=IFERROR(INDEX(Lotes!$E:$E,MATCH(1,(Lotes!$A:$A=A{r})*(Lotes!$C:$C=C{r}),0)),"")'
    ws_stock[f"E{r}"] = f'=IF(D{r}="","",D{r}-TODAY())'
    ws_stock[f"F{r}"] = f'=IF(D{r}="","",IF(E{r}<0,"VENCIDO",IF(E{r}<=30,"VENCE EM <= 30 DIAS",IF(E{r}<=60,"ATENCAO (31-60 DIAS)","OK"))))'
    ws_stock[f"G{r}"] = f'=SUMIFS(Lancamentos!$G:$G,Lancamentos!$C:$C,$A{r},Lancamentos!$E:$E,$C{r},Lancamentos!$B:$B,"ENTRADA")'
    ws_stock[f"H{r}"] = f'=SUMIFS(Lancamentos!$G:$G,Lancamentos!$C:$C,$A{r},Lancamentos!$E:$E,$C{r},Lancamentos!$B:$B,"SAIDA_USO")'
    ws_stock[f"I{r}"] = f'=SUMIFS(Lancamentos!$G:$G,Lancamentos!$C:$C,$A{r},Lancamentos!$E:$E,$C{r},Lancamentos!$B:$B,"SAIDA_DESCARTE")'
    ws_stock[f"J{r}"] = f'=G{r}-H{r}-I{r}'
    ws_stock[f"K{r}"] = f'=IF(A{r}="","",IF(J{r}<=IFERROR(VLOOKUP(A{r},Produtos!$A:$E,5,FALSE),0),"ABAIXO DO MINIMO","OK"))'

for c in ["D"]:
    for r in range(7, 8001):
        ws_stock[f"{c}{r}"].number_format = "dd/mm/yyyy"
for c in ["E", "G", "H", "I", "J"]:
    for r in range(7, 8001):
        ws_stock[f"{c}{r}"].number_format = "0.00"

red = FormulaRule(formula=['$F7="VENCIDO"'], fill=PatternFill("solid", fgColor="FFC7CE"))
yel = FormulaRule(formula=['$F7="VENCE EM <= 30 DIAS"'], fill=PatternFill("solid", fgColor="FFEB9C"))
org = FormulaRule(formula=['$F7="ATENCAO (31-60 DIAS)"'], fill=PatternFill("solid", fgColor="FCE4D6"))
minr = FormulaRule(formula=['$K7="ABAIXO DO MINIMO"'], fill=PatternFill("solid", fgColor="F4B183"))

ws_stock.conditional_formatting.add("A7:K8000", red)
ws_stock.conditional_formatting.add("A7:K8000", yel)
ws_stock.conditional_formatting.add("A7:K8000", org)
ws_stock.conditional_formatting.add("A7:K8000", minr)

table_border(ws_stock, 7, 300, 11)
ws_stock.freeze_panes = "A7"
ws_stock.auto_filter.ref = "A6:K8000"
colw(ws_stock, {"A":14,"B":28,"C":14,"D":12,"E":10,"F":23,"G":12,"H":12,"I":14,"J":12,"K":18})

# RELATORIO
ws_rel.merge_cells("A1:H1")
ws_rel["A1"] = "RELATORIO RAPIDO"
ws_rel["A1"].fill = c_titulo
ws_rel["A1"].font = f_titulo
ws_rel["A1"].alignment = ctr

ws_rel["A3"] = "Indicador"
ws_rel["B3"] = "Valor"
for c in ["A3", "B3"]:
    ws_rel[c].fill = c_header
    ws_rel[c].font = f_head
    ws_rel[c].alignment = ctr
    ws_rel[c].border = b

inds = [
    ("Produtos cadastrados", "=COUNTA(Produtos!A7:A3000)"),
    ("Lotes ativos", "=COUNTA(Estoque_Alertas!C7:C8000)"),
    ("Saldo total", "=SUM(Estoque_Alertas!J7:J8000)"),
    ("Lotes vencidos", '=COUNTIF(Estoque_Alertas!F7:F8000,"VENCIDO")'),
    ("Vence <=30 dias", '=COUNTIF(Estoque_Alertas!F7:F8000,"VENCE EM <= 30 DIAS")'),
    ("Abaixo do minimo", '=COUNTIF(Estoque_Alertas!K7:K8000,"ABAIXO DO MINIMO")'),
]
start = 4
for i, (k, v) in enumerate(inds, start):
    ws_rel[f"A{i}"] = k
    ws_rel[f"B{i}"] = v
    ws_rel[f"A{i}"].border = b
    ws_rel[f"B{i}"].border = b

colw(ws_rel, {"A":30, "B":20, "C":20, "D":20, "E":20, "F":20, "G":20, "H":20})

# COMO USAR
ws_help.merge_cells("A1:H1")
ws_help["A1"] = "INSTRUCOES DE USO (SEM INSTALACAO)"
ws_help["A1"].fill = c_titulo
ws_help["A1"].font = f_titulo
ws_help["A1"].alignment = ctr

txt = [
    "1) Abra a aba Inicio e siga os botoes numerados (1 ate 5).",
    "2) Produtos: cadastrar ID unico, nome, categoria, unidade e estoque minimo.",
    "3) Lotes: cadastrar cada lote com validade correta.",
    "4) Lancamentos: registrar todas as entradas e saidas (uso/descarte).",
    "5) Estoque_Alertas: verificar itens vencidos, proximos do vencimento e abaixo do minimo.",
    "6) Nunca apague formulas nas colunas calculadas.",
    "7) Para evitar conflito em rede: apenas 1 pessoa editando por vez.",
    "8) Salvar diariamente uma copia de backup com data no nome do arquivo.",
    "9) Em caso de erro de digitacao, corrigir no cadastro/lancamento original.",
    "10) Recomendado: filtro por Status_Validade para priorizar acao imediata.",
]

ws_help["A3"] = "Fluxo operacional"
ws_help["A3"].font = f_b
for i, t in enumerate(txt, 5):
    ws_help[f"A{i}"] = t
    ws_help[f"A{i}"].alignment = Alignment(vertical="top", wrap_text=True)

ws_help["A18"] = "Observacao: esta versao funciona apenas com recursos nativos do Excel (sem macros)."
ws_help["A18"].font = Font(italic=True)
colw(ws_help, {"A":120,"B":20,"C":20,"D":20,"E":20,"F":20,"G":20,"H":20})

wb.save(ARQ)
print(f"Arquivo criado: {ARQ}")
