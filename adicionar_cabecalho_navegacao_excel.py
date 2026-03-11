from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ARQ = "Sistema_Estoque_Inflamaveis_SEM_INSTALACAO.xlsx"

wb = load_workbook(ARQ)
sheets = wb.sheetnames

fill_header = PatternFill("solid", fgColor="0B2F44")
fill_info = PatternFill("solid", fgColor="EAF4FB")
fill_btn = PatternFill("solid", fgColor="1F9D67")
font_header = Font(color="FFFFFF", bold=True, size=11)
font_info = Font(color="163447", bold=True, size=9)
font_btn = Font(color="FFFFFF", bold=True, size=10)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
border = Border(
    left=Side(style="thin", color="BFD3E2"),
    right=Side(style="thin", color="BFD3E2"),
    top=Side(style="thin", color="BFD3E2"),
    bottom=Side(style="thin", color="BFD3E2"),
)

hoje = datetime.now().strftime("%d/%m/%Y %H:%M")

for i, nome in enumerate(sheets):
    ws = wb[nome]

    # Cabeçalho visual no topo direito (não interfere na estrutura existente)
    ws.merge_cells("L1:P1")
    c = ws["L1"]
    c.value = "CONTROLE DE ESTOQUE - TINTAS E INFLAMAVEIS"
    c.fill = fill_header
    c.font = font_header
    c.alignment = align_center

    ws.merge_cells("L2:N2")
    ws.merge_cells("O2:P2")

    c1 = ws["L2"]
    c1.value = f"Aba atual: {nome}"
    c1.fill = fill_info
    c1.font = font_info
    c1.alignment = align_left

    c2 = ws["O2"]
    c2.value = f"Atualizado: {hoje}"
    c2.fill = fill_info
    c2.font = font_info
    c2.alignment = align_center

    # Navegação
    prev_sheet = sheets[i - 1] if i > 0 else sheets[-1]
    next_sheet = sheets[i + 1] if i < len(sheets) - 1 else sheets[0]

    ws.merge_cells("L3:M3")
    ws.merge_cells("N3:O3")
    ws["P3"] = ""

    b_prev = ws["L3"]
    b_prev.value = "<< Anterior"
    b_prev.hyperlink = f"#'{prev_sheet}'!A1"
    b_prev.fill = fill_btn
    b_prev.font = font_btn
    b_prev.alignment = align_center

    b_home = ws["N3"]
    b_home.value = "Inicio"
    b_home.hyperlink = "#'Inicio'!A1"
    b_home.fill = fill_btn
    b_home.font = font_btn
    b_home.alignment = align_center

    b_next = ws["P3"]
    b_next.value = "Proxima >>"
    b_next.hyperlink = f"#'{next_sheet}'!A1"
    b_next.fill = fill_btn
    b_next.font = font_btn
    b_next.alignment = align_center

    for rng in ["L1:P1", "L2:P2", "L3:P3"]:
        for row in ws[rng]:
            for cell in row:
                cell.border = border

    # Larguras das colunas de cabeçalho auxiliar
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["M"].width = 15
    ws.column_dimensions["N"].width = 15
    ws.column_dimensions["O"].width = 15
    ws.column_dimensions["P"].width = 15

wb.save(ARQ)
print(f"Cabecalho aplicado em todas as abas: {ARQ}")
